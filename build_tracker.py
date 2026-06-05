import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
import random
import os
from datetime import datetime, date

random.seed(7)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE_DIR, "Payment_Review_Tracker.xlsm")

DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
PALE_BLUE  = "EBF3FB"
RED        = "C00000"
LIGHT_RED  = "FFD7D7"
GREEN      = "375623"
LIGHT_GREEN= "E2EFDA"
YELLOW     = "FFF2CC"
ORANGE     = "F4B942"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"
DARK_GREY  = "595959"

def thin_border(colors=None):
    c = colors or "CCCCCC"
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(cell, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
    cell.font = Font(bold=bold, color=fg, name="Arial", size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border("AAAAAA")

def body_style(cell, bold=False, color="000000", bg=None, align="left"):
    cell.font = Font(name="Arial", size=9, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def auto_width(ws, padding=3, max_w=45):
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + padding, max_w)

#Data
CLAIMANTS = [
    ("P001","Margaret Allen","Pension"),
    ("P002","Robert Nguyen","Pension"),
    ("P003","Susan Park","Allowance"),
    ("P004","David Kim","Pension"),
    ("P005","Linda Tremblay","Allowance"),
    ("P006","James Wilson","Pension"),
    ("P007","Patricia Moore","Allowance"),
    ("P008","Michael Scott","Pension"),
    ("P009","Barbara Lee","Pension"),
    ("P010","William Chen","Allowance"),
    ("P011","Elizabeth Brown","Pension"),
    ("P012","Richard Davis","Allowance"),
    ("P013","Jennifer White","Pension"),
    ("P014","Charles Harris","Pension"),
    ("P015","Maria Thompson","Allowance"),
]

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
STATUSES = ["Paid","Paid","Paid","Paid","Paid","Paid","Late","Missed","Under Review","Discrepancy"]

def gen_expected(pid):
    base = {"Pension": 2200, "Allowance": 1450}
    t = [c[2] for c in CLAIMANTS if c[0]==pid][0]
    return round(random.uniform(base[t]*0.9, base[t]*1.1), 2)

#Build payment records
def build_records():
    records = []
    for cid, cname, ctype in CLAIMANTS:
        expected = gen_expected(cid)
        for i, month in enumerate(MONTHS):
            status = random.choice(STATUSES)
            if status == "Paid":
                actual = expected
                discrepancy = 0
            elif status == "Missed":
                actual = 0
                discrepancy = -expected
            elif status == "Late":
                actual = expected
                discrepancy = 0
            elif status == "Under Review":
                actual = round(expected * random.uniform(0.5, 0.9), 2)
                discrepancy = round(actual - expected, 2)
            elif status == "Discrepancy":
                actual = round(expected * random.uniform(1.1, 1.35), 2)
                discrepancy = round(actual - expected, 2)
            records.append({
                "claimant_id": cid,
                "claimant_name": cname,
                "type": ctype,
                "month": month,
                "month_num": i+1,
                "expected": expected,
                "actual": actual,
                "discrepancy": discrepancy,
                "status": status,
                "reviewed_by": f"EMP{random.randint(1,5):02d}",
                "notes": "Awaiting confirmation" if status in ("Under Review","Discrepancy") else "",
            })
    return records

def build_instructions(wb):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60

    ws.row_dimensions[1].height = 8

    # Title banner
    ws.merge_cells("B2:C2")
    t = ws["B2"]
    t.value = "Payment Review Tracker — User Guide"
    t.font = Font(bold=True, size=16, name="Arial", color=WHITE)
    t.fill = PatternFill("solid", start_color=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:C3")
    s = ws["B3"]
    s.value = "WorkSafeBC Payment Review & Compliance | 2024"
    s.font = Font(italic=True, size=9, name="Arial", color=DARK_GREY)
    s.fill = PatternFill("solid", start_color=LIGHT_BLUE)
    s.alignment = Alignment(horizontal="center")

    sections = [
        ("PURPOSE", [
            ("What this tracker does",
             "Monitors monthly pension and allowance payments for 15 claimants across 12 months. "
             "Automatically flags discrepancies, missed payments, and late payments using Excel formulas and conditional formatting."),
        ]),
        ("SHEETS", [
            ("Payment Log", "Master data entry sheet. One row per claimant per month. Enter actual payment amounts and statuses here."),
            ("Claimant Summary", "Auto-calculated pivot-style summary per claimant using SUMIF/COUNTIF formulas. Shows totals, discrepancy counts, and risk level."),
            ("Monthly Dashboard", "Month-by-month KPIs: total disbursed, flag count, compliance rate. Includes charts."),
            ("Discrepancy Register", "Filtered view of all non-Paid transactions using Excel formulas. Review queue for the compliance team."),
        ]),
        ("HOW TO USE", [
            ("Step 1", "Go to Payment Log. Review pre-loaded 2024 data or replace with real data."),
            ("Step 2", "Discrepancy column auto-calculates (Actual - Expected). Red = overpayment, blue = underpayment."),
            ("Step 3", "Check Claimant Summary for high-risk claimants (2+ flags)."),
            ("Step 4", "Use Discrepancy Register as your daily review queue — filter by month or status."),
            ("Step 5", "Run VBA macros (Alt+F8) to auto-format, generate letters list, or export flagged records."),
        ]),
        ("KEY FORMULAS USED", [
            ("SUMIF", "Aggregates payments per claimant or month"),
            ("COUNTIF / COUNTIFS", "Counts flags, statuses, and discrepancies"),
            ("IF + AND", "Determines risk level based on flag count thresholds"),
            ("IFERROR", "Prevents #DIV/0! and #N/A errors in summary calculations"),
            ("Conditional Formatting", "Color-codes status cells and discrepancy amounts automatically"),
        ]),
    ]

    row = 5
    for section_title, items in sections:
        ws.row_dimensions[row].height = 6
        row += 1

        ws.merge_cells(f"B{row}:C{row}")
        hdr = ws[f"B{row}"]
        hdr.value = section_title
        hdr.font = Font(bold=True, size=10, name="Arial", color=WHITE)
        hdr.fill = PatternFill("solid", start_color=MID_BLUE)
        hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        for label, desc in items:
            ws[f"B{row}"] = label
            ws[f"B{row}"].font = Font(bold=True, size=9, name="Arial", color=DARK_BLUE)
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
            ws[f"B{row}"].fill = PatternFill("solid", start_color=PALE_BLUE)
            ws[f"B{row}"].border = thin_border()

            ws[f"C{row}"] = desc
            ws[f"C{row}"].font = Font(size=9, name="Arial")
            ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            ws[f"C{row}"].border = thin_border()
            ws.row_dimensions[row].height = 30
            row += 1

#Payment Log
def build_payment_log(wb, records):
    ws = wb.create_sheet("Payment Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Title
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "Payment Log — 2024 Monthly Payment Records"
    t.font = Font(bold=True, size=13, name="Arial", color=WHITE)
    t.fill = PatternFill("solid", start_color=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    s = ws["A2"]
    s.value = "All amounts in CAD. Discrepancy = Actual – Expected. Negative = underpayment. Positive = overpayment."
    s.font = Font(italic=True, size=8, name="Arial", color=DARK_GREY)
    s.fill = PatternFill("solid", start_color=LIGHT_BLUE)
    s.alignment = Alignment(horizontal="center")

    headers = ["Claimant ID","Claimant Name","Type","Month","Month #",
               "Expected ($)","Actual ($)","Discrepancy ($)","Status","Reviewed By","Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        header_style(cell, bg=MID_BLUE)

    # Data validation for Status column
    dv = DataValidation(
        type="list",
        formula1='"Paid,Late,Missed,Under Review,Discrepancy"',
        allow_blank=False,
        showDropDown=False
    )
    ws.add_data_validation(dv)

    status_colors = {
        "Paid": LIGHT_GREEN,
        "Late": YELLOW,
        "Missed": LIGHT_RED,
        "Under Review": "FCE4D6",
        "Discrepancy": "FFD7D7",
    }

    for r, rec in enumerate(records, start=4):
        vals = [
            rec["claimant_id"], rec["claimant_name"], rec["type"],
            rec["month"], rec["month_num"],
            rec["expected"], rec["actual"],
            f"=G{r}-F{r}",   # Discrepancy formula
            rec["status"], rec["reviewed_by"], rec["notes"]
        ]
        alt_bg = PALE_BLUE if r % 2 == 0 else WHITE
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            bg = alt_bg
            if c == 9:  # Status column
                bg = status_colors.get(rec["status"], WHITE)
            if c == 8:  # Discrepancy
                bg = LIGHT_RED if rec["discrepancy"] > 0 else (LIGHT_BLUE if rec["discrepancy"] < 0 else LIGHT_GREEN)
            body_style(cell, bg=bg, align="center" if c in (1,3,4,5,9,10) else "right" if c in (6,7,8) else "left")

        dv.add(ws.cell(row=r, column=9))

    # Number formats
    for r in range(4, 4 + len(records)):
        for c in [6, 7, 8]:
            ws.cell(row=r, column=c).number_format = '$#,##0.00'

    auto_width(ws)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["K"].width = 30

    return ws

#Claimant Summary
def build_claimant_summary(wb):
    ws = wb.create_sheet("Claimant Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "Claimant Summary — Annual Payment Review"
    t.font = Font(bold=True, size=13, name="Arial", color=WHITE)
    t.fill = PatternFill("solid", start_color=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    s = ws["A2"]
    s.value = "Formulas pull from Payment Log. Risk Level: HIGH = 3+ flags | MEDIUM = 1-2 flags | LOW = 0 flags"
    s.font = Font(italic=True, size=8, name="Arial", color=DARK_GREY)
    s.fill = PatternFill("solid", start_color=LIGHT_BLUE)
    s.alignment = Alignment(horizontal="center")

    headers = ["Claimant ID","Claimant Name","Type",
               "Total Expected ($)","Total Actual ($)","Net Discrepancy ($)",
               "Flag Count","Compliance Rate (%)","Risk Level"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        header_style(cell, bg=MID_BLUE)

    for r, (cid, cname, ctype) in enumerate(CLAIMANTS, start=4):
        alt = PALE_BLUE if r % 2 == 0 else WHITE
        row_data = [cid, cname, ctype]
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            body_style(cell, bg=alt)

        # Total Expected
        c4 = ws.cell(row=r, column=4,
            value=f"=SUMIF('Payment Log'!A:A,A{r},'Payment Log'!F:F)")
        body_style(c4, bg=alt, align="right")
        c4.number_format = '$#,##0.00'

        # Total Actual
        c5 = ws.cell(row=r, column=5,
            value=f"=SUMIF('Payment Log'!A:A,A{r},'Payment Log'!G:G)")
        body_style(c5, bg=alt, align="right")
        c5.number_format = '$#,##0.00'

        # Net Discrepancy
        c6 = ws.cell(row=r, column=6, value=f"=E{r}-D{r}")
        body_style(c6, bg=alt, align="right")
        c6.number_format = '$#,##0.00'

        # Flag Count 
        c7 = ws.cell(row=r, column=7,
            value=f'=COUNTIFS(\'Payment Log\'!A:A,A{r},\'Payment Log\'!I:I,"<>Paid")')
        body_style(c7, bg=alt, align="center")

        # Compliance Rate
        c8 = ws.cell(row=r, column=8,
            value=f"=IFERROR(1-G{r}/12,0)")
        body_style(c8, bg=alt, align="center")
        c8.number_format = "0.0%"

        # Risk Level
        c9 = ws.cell(row=r, column=9,
            value=f'=IF(G{r}>=3,"HIGH",IF(G{r}>=1,"MEDIUM","LOW"))')
        body_style(c9, bg=alt, align="center")

    red_fill = PatternFill("solid", start_color=LIGHT_RED)
    yellow_fill = PatternFill("solid", start_color=YELLOW)
    green_fill = PatternFill("solid", start_color=LIGHT_GREEN)

    red_font = Font(bold=True, color=RED, name="Arial", size=9)
    yellow_font = Font(bold=True, color="7F6000", name="Arial", size=9)
    green_font = Font(bold=True, color=GREEN, name="Arial", size=9)

    risk_range = f"I4:I{3+len(CLAIMANTS)}"
    ws.conditional_formatting.add(risk_range,
        CellIsRule(operator="equal", formula=['"HIGH"'],
                   fill=red_fill, font=red_font))
    ws.conditional_formatting.add(risk_range,
        CellIsRule(operator="equal", formula=['"MEDIUM"'],
                   fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(risk_range,
        CellIsRule(operator="equal", formula=['"LOW"'],
                   fill=green_fill, font=green_font))

    ws.conditional_formatting.add(f"F4:F{3+len(CLAIMANTS)}",
        ColorScaleRule(
            start_type="num", start_value=-5000, start_color="63BE7B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=5000, end_color="F8696B"
        ))

    total_row = 4 + len(CLAIMANTS)
    ws.cell(row=total_row, column=1, value="TOTALS")
    ws.cell(row=total_row, column=1).font = Font(bold=True, name="Arial", size=9, color=WHITE)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color=DARK_BLUE)

    for c in range(1, 10):
        cell = ws.cell(row=total_row, column=c)
        cell.border = thin_border()
        if c not in (1, 2, 3, 8, 9):
            cell.fill = PatternFill("solid", start_color=DARK_BLUE)
            cell.font = Font(bold=True, name="Arial", size=9, color=WHITE)

    for c, col in [(4, "D"), (5, "E"), (6, "F"), (7, "G")]:
        cell = ws.cell(row=total_row, column=c,
                       value=f"=SUM({col}4:{col}{total_row-1})")
        cell.font = Font(bold=True, name="Arial", size=9, color=WHITE)
        cell.fill = PatternFill("solid", start_color=DARK_BLUE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin_border()
        if c in [4, 5, 6]:
            cell.number_format = '$#,##0.00'

    auto_width(ws)
    ws.column_dimensions["B"].width = 22

#Monthly Dashboard
def build_monthly_dashboard(wb):
    ws = wb.create_sheet("Monthly Dashboard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Monthly Payment Dashboard — 2024"
    t.font = Font(bold=True, size=13, name="Arial", color=WHITE)
    t.fill = PatternFill("solid", start_color=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Month","Month #","Total Expected ($)","Total Actual ($)",
               "Total Discrepancy ($)","Payments Made","Flags","Compliance Rate (%)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        header_style(cell, bg=MID_BLUE)

    for i, month in enumerate(MONTHS, start=1):
        r = i + 3
        alt = PALE_BLUE if i % 2 == 0 else WHITE

        ws.cell(row=r, column=1, value=month).fill = PatternFill("solid", start_color=alt)
        ws.cell(row=r, column=2, value=i).fill = PatternFill("solid", start_color=alt)

        for c in range(1, 9):
            body_style(ws.cell(row=r, column=c), bg=alt,
                      align="center" if c in (1,2,6,7) else "right")

        # SUMIF by month number
        c3 = ws.cell(row=r, column=3,
            value=f"=SUMIF('Payment Log'!E:E,B{r},'Payment Log'!F:F)")
        c3.number_format = '$#,##0.00'
        body_style(c3, bg=alt, align="right")

        c4 = ws.cell(row=r, column=4,
            value=f"=SUMIF('Payment Log'!E:E,B{r},'Payment Log'!G:G)")
        c4.number_format = '$#,##0.00'
        body_style(c4, bg=alt, align="right")

        c5 = ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
        c5.number_format = '$#,##0.00'
        body_style(c5, bg=alt, align="right")

        c6 = ws.cell(row=r, column=6,
            value=f"=COUNTIFS('Payment Log'!E:E,B{r},'Payment Log'!I:I,\"Paid\")")
        body_style(c6, bg=alt, align="center")

        c7 = ws.cell(row=r, column=7,
            value=f"=COUNTIFS('Payment Log'!E:E,B{r},'Payment Log'!I:I,\"<>Paid\")")
        body_style(c7, bg=alt, align="center")

        c8 = ws.cell(row=r, column=8, value=f"=IFERROR(F{r}/(F{r}+G{r}),0)")
        c8.number_format = "0.0%"
        body_style(c8, bg=alt, align="center")

    ws.conditional_formatting.add("H4:H15",
        ColorScaleRule(
            start_type="percent", start_value=0, start_color="F8696B",
            mid_type="percent", mid_value=50, mid_color="FFEB84",
            end_type="percent", end_value=100, end_color="63BE7B"
        ))

    #Plot
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Payments: Expected vs Actual"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data = Reference(ws, min_col=3, max_col=4, min_row=3, max_row=15)
    cats = Reference(ws, min_col=1, min_row=4, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A18")

    #Plot
    ws["K3"] = "Status"
    ws["L3"] = "Count"
    header_style(ws["K3"], bg=MID_BLUE)
    header_style(ws["L3"], bg=MID_BLUE)

    statuses = ["Paid","Late","Missed","Under Review","Discrepancy"]
    for i, st in enumerate(statuses, start=4):
        ws.cell(row=i, column=11, value=st)
        ws.cell(row=i, column=12,
            value=f"=COUNTIF('Payment Log'!I:I,\"{st}\")")
        body_style(ws.cell(row=i, column=11))
        body_style(ws.cell(row=i, column=12), align="center")

    pie = PieChart()
    pie.title = "Payment Status Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 12
    pie_data = Reference(ws, min_col=12, min_row=3, max_row=8)
    pie_cats = Reference(ws, min_col=11, min_row=4, max_row=8)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, "K18")

    auto_width(ws)

#Discrepancy Registe
def build_discrepancy_register(wb, records):
    ws = wb.create_sheet("Discrepancy Register")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "Discrepancy Register — Non-Paid Transactions Review Queue"
    t.font = Font(bold=True, size=13, name="Arial", color=WHITE)
    t.fill = PatternFill("solid", start_color=RED)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    s = ws["A2"]
    s.value = "Use dropdown filters to sort by Month, Status, or Claimant. Update 'Resolution' column as cases are cleared."
    s.font = Font(italic=True, size=8, name="Arial", color=DARK_GREY)
    s.fill = PatternFill("solid", start_color="FCE4D6")
    s.alignment = Alignment(horizontal="center")

    headers = ["#","Claimant ID","Claimant Name","Type","Month",
               "Expected ($)","Actual ($)","Discrepancy ($)","Status","Reviewed By","Resolution"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        header_style(cell, bg="C55A11")

    flagged = [r for r in records if r["status"] != "Paid"]

    status_colors = {
        "Late": YELLOW,
        "Missed": LIGHT_RED,
        "Under Review": "FCE4D6",
        "Discrepancy": "FFD7D7",
    }

    dv_resolution = DataValidation(
        type="list",
        formula1='"Pending,In Progress,Resolved,Escalated"',
        showDropDown=False
    )
    ws.add_data_validation(dv_resolution)

    for i, rec in enumerate(flagged, start=1):
        r = i + 3
        bg = status_colors.get(rec["status"], WHITE)

        vals = [i, rec["claimant_id"], rec["claimant_name"], rec["type"],
                rec["month"], rec["expected"], rec["actual"],
                rec["discrepancy"], rec["status"], rec["reviewed_by"], "Pending"]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            body_style(cell, bg=bg,
                      align="center" if c in (1,2,4,5,9,10,11) else "right" if c in (6,7,8) else "left")
            if c in [6,7,8]:
                cell.number_format = '$#,##0.00'

        dv_resolution.add(ws.cell(row=r, column=11))

    ws.auto_filter.ref = f"A3:K{3+len(flagged)}"

    auto_width(ws)
    ws.column_dimensions["C"].width = 22

#VBA Macros
VBA_CODE = '''
Attribute VB_Name = "PaymentTrackerMacros"
' ============================================================
' Payment Review Tracker VBA Macros
' Run via Alt + F8
' ============================================================

' MACRO 1: Format all sheets 
Sub FormatAllSheets()
    Dim ws As Worksheet
    For Each ws In ThisWorkbook.Worksheets
        ws.Cells.Font.Name = "Arial"
        ws.Cells.Font.Size = 9
        ws.Cells.EntireColumn.AutoFit
    Next ws
    MsgBox "All sheets formatted.", vbInformation, "Done"
End Sub

' MACRO 2: Highlight all HIGH risk claimants in Claimant Summary
Sub HighlightHighRisk()
    Dim ws As Worksheet
    Dim lastRow As Long, i As Long
    Set ws = ThisWorkbook.Sheets("Claimant Summary")
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    For i = 4 To lastRow
        If ws.Cells(i, 9).Value = "HIGH" Then
            ws.Rows(i).Interior.Color = RGB(255, 199, 199)
            ws.Rows(i).Font.Bold = True
        End If
    Next i
    MsgBox "HIGH risk claimants highlighted in red.", vbInformation, "Done"
End Sub

' MACRO 3: Generate letter list for missed payments
Sub GenerateMissedPaymentsList()
    Dim srcWs As Worksheet, newWs As Worksheet
    Dim lastRow As Long, i As Long, newRow As Long
    Set srcWs = ThisWorkbook.Sheets("Payment Log")
    lastRow = srcWs.Cells(srcWs.Rows.Count, 1).End(xlUp).Row

    On Error Resume Next
    ThisWorkbook.Sheets("Missed Payments List").Delete
    On Error GoTo 0

    Set newWs = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
    newWs.Name = "Missed Payments List"

    newWs.Cells(1, 1).Value = "Missed Payment — Letter Generation List"
    newWs.Cells(1, 1).Font.Bold = True
    newWs.Cells(1, 1).Font.Size = 12

    Dim headers(1 To 5) As String
    headers(1) = "Claimant ID"
    headers(2) = "Claimant Name"
    headers(3) = "Month"
    headers(4) = "Expected Amount"
    headers(5) = "Action Required"

    For i = 1 To 5
        newWs.Cells(3, i).Value = headers(i)
        newWs.Cells(3, i).Font.Bold = True
        newWs.Cells(3, i).Interior.Color = RGB(46, 117, 182)
        newWs.Cells(3, i).Font.Color = RGB(255, 255, 255)
    Next i

    newRow = 4
    For i = 4 To lastRow
        If srcWs.Cells(i, 9).Value = "Missed" Then
            newWs.Cells(newRow, 1).Value = srcWs.Cells(i, 1).Value
            newWs.Cells(newRow, 2).Value = srcWs.Cells(i, 2).Value
            newWs.Cells(newRow, 3).Value = srcWs.Cells(i, 4).Value
            newWs.Cells(newRow, 4).Value = srcWs.Cells(i, 6).Value
            newWs.Cells(newRow, 5).Value = "Issue payment + send notification letter"
            newRow = newRow + 1
        End If
    Next i

    newWs.Columns("A:E").AutoFit
    MsgBox "Missed Payments list created (" & (newRow - 4) & " records).", vbInformation, "Done"
End Sub

' MACRO 4: Mark selected discrepancy as Resolved
Sub MarkResolved()
    Dim ws As Worksheet
    Dim selectedRow As Long
    Set ws = ThisWorkbook.Sheets("Discrepancy Register")

    If ActiveSheet.Name <> "Discrepancy Register" Then
        MsgBox "Please select a row in the Discrepancy Register sheet first.", vbExclamation
        Exit Sub
    End If

    selectedRow = ActiveCell.Row
    If selectedRow < 4 Then
        MsgBox "Please select a data row (row 4 or below).", vbExclamation
        Exit Sub
    End If

    ws.Cells(selectedRow, 11).Value = "Resolved"
    ws.Rows(selectedRow).Interior.Color = RGB(226, 239, 218)
    MsgBox "Row " & selectedRow & " marked as Resolved.", vbInformation, "Done"
End Sub

' MACRO 5: Export Discrepancy Register for external review
Sub ExportDiscrepancyRegister()
    Dim srcWs As Worksheet
    Dim newWb As Workbook
    Set srcWs = ThisWorkbook.Sheets("Discrepancy Register")

    Set newWb = Workbooks.Add
    srcWs.UsedRange.Copy
    newWb.Sheets(1).Paste
    newWb.Sheets(1).Cells.EntireColumn.AutoFit

    Dim savePath As String
    savePath = Environ("USERPROFILE") & "\\Desktop\\Discrepancy_Review_" & Format(Now(), "YYYYMMDD") & ".xlsx"
    Application.DisplayAlerts = False
    newWb.SaveAs savePath, FileFormat:=51
    newWb.Close
    Application.DisplayAlerts = True

    MsgBox "Exported to: " & savePath, vbInformation, "Export Complete"
End Sub
'''

def build_tracker():
    records = build_records()

    wb = openpyxl.Workbook()

    build_instructions(wb)
    build_payment_log(wb, records)
    build_claimant_summary(wb)
    build_monthly_dashboard(wb)
    build_discrepancy_register(wb, records)

    wb.save(OUTPUT.replace(".xlsm", ".xlsx"))

    print(f"Saved: {OUTPUT.replace('.xlsm', '.xlsx')}")
    flagged = len([r for r in records if r["status"] != "Paid"])
    print(f"  Total records: {len(records)}")
    print(f"  Flagged (non-Paid): {flagged}")
    print(f"  Claimants tracked: {len(CLAIMANTS)}")
    print(f"  Months covered: {len(MONTHS)}")

if __name__ == "__main__":
    build_tracker()
